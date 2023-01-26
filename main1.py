from itertools import product
import openpyxl

def main():


    inv_file = openpyxl.load_workbook("inventory.xlsx")
    product_list =  inv_file["Sheet1"]



    # create a dictionary
    product_per_supplier = {}
    total_value_per_supplier = {}
    products_under_10_inv = {}

    #print(product_list.max_row)

    for products_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(products_row, 4).value
        inventory = product_list.cell(products_row, 2).value 
        price = product_list.cell(products_row, 3).value
        product_number = product_list.cell(products_row, 1).value
        inventory_price = product_list.cell(products_row, 5)


        # calculation number of products per supplier.
        if supplier_name in product_per_supplier:
            current_num_products =  product_per_supplier.get(supplier_name)
            product_per_supplier[supplier_name] = current_num_products + 1 
            

        else:

          
            product_per_supplier[supplier_name] = 1

        # calculation total value of inventory per supplier  
        if supplier_name in total_value_per_supplier:
           current_total_value  = total_value_per_supplier.get(supplier_name)
           total_value_per_supplier[supplier_name] = current_total_value + inventory * price
          
        else:
            
          total_value_per_supplier[supplier_name] = inventory * price 
    
         # logiv products with inventory lees than 10.   
        if inventory < 10:
           products_under_10_inv[product_number] = inventory

        # add value for total inventory price:
        inventory_price.value = inventory * price


    print()
    print(product_per_supplier)
    print()
    print(total_value_per_supplier)
    print()
    print(products_under_10_inv)
    input("Press <Enter> to exit!!")

    inv_file.save("inventory_with_total_value.xlsx")

main()

     
